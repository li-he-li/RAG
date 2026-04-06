"""
Shared upload text extraction helpers for document-like files.
"""

from __future__ import annotations

import io
import logging
import re
import shutil
import subprocess
import tempfile
from pathlib import Path

from fastapi import HTTPException


logger = logging.getLogger(__name__)
DOC_PREVIEW_TEXT_RE = re.compile(r"[\u4e00-\u9fffA-Za-z0-9][^\x00]{3,}")


def _normalize_extracted_text(text: str) -> str:
    lines = [line.strip() for line in text.replace("\r\n", "\n").replace("\r", "\n").split("\n")]
    cleaned: list[str] = []
    previous = None
    for line in lines:
        compact = re.sub(r"\s+", " ", line).strip()
        if not compact or compact == previous:
            continue
        cleaned.append(compact)
        previous = compact
    return "\n".join(cleaned).strip()


def _ensure_non_empty_text(text: str, detail: str) -> str:
    normalized = _normalize_extracted_text(text)
    if not normalized:
        raise HTTPException(status_code=422, detail=detail)
    return normalized


def _extract_xlsx_text(raw: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"XLSX parser unavailable: {exc}") from exc

    workbook = load_workbook(filename=io.BytesIO(raw), data_only=True, read_only=True)
    rows: list[str] = []
    for sheet in workbook.worksheets:
        rows.append(f"[工作表] {sheet.title}")
        for values in sheet.iter_rows(values_only=True):
            cells = [str(value).strip() for value in values if value is not None and str(value).strip()]
            if cells:
                rows.append(" | ".join(cells))
    return _ensure_non_empty_text("\n".join(rows), "Excel 文件未提取到可用文本内容")


def _extract_xls_text(raw: bytes) -> str:
    try:
        import xlrd
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"XLS parser unavailable: {exc}") from exc

    workbook = xlrd.open_workbook(file_contents=raw)
    rows: list[str] = []
    for sheet in workbook.sheets():
        rows.append(f"[工作表] {sheet.name}")
        for row_idx in range(sheet.nrows):
            values: list[str] = []
            for col_idx in range(sheet.ncols):
                value = sheet.cell_value(row_idx, col_idx)
                if value is None:
                    continue
                text = str(value).strip()
                if text:
                    values.append(text)
            if values:
                rows.append(" | ".join(values))
    return _ensure_non_empty_text("\n".join(rows), "Excel 文件未提取到可用文本内容")


def _extract_doc_text_via_word(raw: bytes) -> str:
    try:
        import pythoncom
        import win32com.client
    except Exception:
        return ""

    with tempfile.TemporaryDirectory() as temp_dir:
        source_path = Path(temp_dir) / "upload.doc"
        output_path = Path(temp_dir) / "upload.txt"
        source_path.write_bytes(raw)

        pythoncom.CoInitialize()
        word = None
        doc = None
        try:
            word = win32com.client.Dispatch("Word.Application")
            word.Visible = False
            word.DisplayAlerts = 0
            doc = word.Documents.Open(str(source_path), ReadOnly=True)
            doc.SaveAs(str(output_path), FileFormat=7)
        finally:
            if doc is not None:
                doc.Close(False)
            if word is not None:
                word.Quit()
            pythoncom.CoUninitialize()

        if not output_path.exists():
            return ""
        return output_path.read_text(encoding="utf-16", errors="ignore")


def _extract_doc_text_via_antiword(raw: bytes) -> str:
    antiword_path = shutil.which("antiword")
    if not antiword_path:
        return ""

    with tempfile.NamedTemporaryFile(suffix=".doc", delete=False) as temp_file:
        temp_path = Path(temp_file.name)
        temp_file.write(raw)

    try:
        completed = subprocess.run(
            [antiword_path, str(temp_path)],
            check=False,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="ignore",
        )
        if completed.returncode != 0:
            return ""
        return completed.stdout
    finally:
        temp_path.unlink(missing_ok=True)


def _extract_doc_stream_candidates(stream: bytes) -> list[str]:
    candidates: list[str] = []
    for encoding, offsets in (("utf-16le", (0, 1)), ("gb18030", (0,)), ("utf-8", (0,)), ("latin1", (0,))):
        for offset in offsets:
            if offset >= len(stream):
                continue
            try:
                decoded = stream[offset:].decode(encoding, errors="ignore")
            except Exception:
                continue
            decoded = decoded.replace("\x00", " ")
            matches = DOC_PREVIEW_TEXT_RE.findall(decoded)
            if matches:
                candidates.append("\n".join(matches))
    return candidates


def _extract_doc_text_via_ole(raw: bytes) -> str:
    try:
        import olefile
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"DOC parser unavailable: {exc}") from exc

    candidates: list[str] = []
    with olefile.OleFileIO(io.BytesIO(raw)) as ole:
        for stream_name in ole.listdir(streams=True, storages=False):
            try:
                stream = ole.openstream(stream_name).read()
            except Exception:
                continue
            candidates.extend(_extract_doc_stream_candidates(stream))

    unique_lines: list[str] = []
    seen = set()
    for candidate in candidates:
        for line in candidate.splitlines():
            compact = re.sub(r"\s+", " ", line).strip()
            if len(compact) < 4 or compact in seen:
                continue
            seen.add(compact)
            unique_lines.append(compact)
    return "\n".join(unique_lines)


def _extract_doc_text(raw: bytes) -> str:
    for extractor in (_extract_doc_text_via_word, _extract_doc_text_via_antiword, _extract_doc_text_via_ole):
        try:
            text = extractor(raw)
        except HTTPException:
            raise
        except Exception:
            logger.debug("DOC extractor failed: %s", extractor.__name__, exc_info=True)
            continue
        if text:
            return _ensure_non_empty_text(text, "DOC 文件未提取到可用文本内容")

    raise HTTPException(status_code=422, detail="DOC 文件未提取到可用文本内容")


def extract_upload_text(file_name: str, raw: bytes) -> str:
    """Extract normalized text from supported upload types."""
    ext = Path(file_name).suffix.lower()

    if ext in {".txt", ".md"}:
        return _ensure_non_empty_text(raw.decode("utf-8", errors="replace"), "文件未提取到可用文本内容")

    if ext == ".pdf":
        try:
            from pypdf import PdfReader
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"PDF parser unavailable: {exc}") from exc

        reader = PdfReader(io.BytesIO(raw))
        text = "\n".join((page.extract_text() or "").strip() for page in reader.pages)
        return _ensure_non_empty_text(text, "PDF 文件未提取到可用文本内容")

    if ext == ".docx":
        try:
            from docx import Document
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"DOCX parser unavailable: {exc}") from exc

        doc = Document(io.BytesIO(raw))
        text = "\n".join((p.text or "").strip() for p in doc.paragraphs)
        return _ensure_non_empty_text(text, "DOCX 文件未提取到可用文本内容")

    if ext == ".doc":
        return _extract_doc_text(raw)

    if ext == ".xlsx":
        return _extract_xlsx_text(raw)

    if ext == ".xls":
        return _extract_xls_text(raw)

    raise HTTPException(
        status_code=415,
        detail="不支持的文件类型，仅支持 .txt .md .pdf .doc .docx .xls .xlsx",
    )
